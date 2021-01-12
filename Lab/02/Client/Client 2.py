from flask import Flask, request, jsonify
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

i = 1
arr = []


def Table(sheet):
    sheet.column_dimensions['A'].width = 15;
    sheet['A1'] = "№"
    sheet.column_dimensions['B'].width = 50;
    sheet['B1'] = "ID"
    sheet.column_dimensions['C'].width = 30;
    sheet['C1'] = "время.дата покупки"
    sheet.column_dimensions['D'].width = 60;
    sheet['D1'] = "Заказ"
    sheet.column_dimensions['E'].width = 20;
    sheet['E1'] = "Цена заказа\n(рублей)"


def Style(sheet, j, x):
    Tst = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    for ij in range(1, j):
        for ix in range(0, x):
            sheet[ij][ix].border = Tst
            sheet[ij][ix].alignment = Alignment(horizontal='center', vertical='center')
    return sheet


def prID(sheet, j):
    i = j - 1
    while (i > 1) and (sheet[i][0].value is None):
        i = i - 1
    if i <= 1:
        return 1
    else:
        return sheet[i][0].value + 1


def Json(sheet):
    j = getId(sheet)
    for i in range(len(arr)):
        sheet[j][0].value = 1 if j <= 1 else prID(sheet, j)
        sheet[j][1].value = arr[i]["user_id"]
        sheet[j][2].value = arr[i]["datetime"]
        firstr = j
        for item in arr[i]["check"]:
            sheet[j][3].value = item["item"]
            sheet[j][4].value = item["price"]
            j = j + 1
        sheet.merge_cells(start_row=firstr, start_column=1, end_row=j - 1, end_column=1)
        sheet.merge_cells(start_row=firstr, start_column=2, end_row=j - 1, end_column=2)
        sheet.merge_cells(start_row=firstr, start_column=3, end_row=j - 1, end_column=3)
    sheet = Style(sheet, j, 5)
    return sheet


def save():
    global arr
    try:
        book = openpyxl.open("data.xlsx", read_only=False)
    except:
        book = openpyxl.Workbook()
    sheet = book.active
    if sheet['A1'].value is None:
        sheet = Table(sheet)
    sheet = Json(sheet)
    book.save("data.xlsx")


def Datatime(content):
    global arr
    arr.append(content);
    arr[-1]["datetime"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    if len(arr) >= i:
        save()
        arr = []


def getId(sheet):
    i = 1
    while not (sheet[i][3].value is None):
        i = i + 1
    return i


app = Flask(__name__)


@app.route('/', methods=['POST', 'GET'])
def index():
    if request.is_json:
        content = request.get_json()
        Datatime(content)
        return 'OK'


app.run()